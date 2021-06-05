from openpyxl.styles.borders import Side
import openpyxl
import requests, random, shutil, json, os, re, sys
from time import sleep
import datetime as dt
from howlongtobeatpy import HowLongToBeat
from bs4 import BeautifulSoup
from tqdm import tqdm


class Indexer:

    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    changes_made = 0


    def __init__(self, excel_filename, workbook_name, column_name, column_letter):
        '''
        Class object init.
        '''
        self.excel_filename = excel_filename
        self.file_path = os.path.join(os.getcwd(), excel_filename + '.xlsx')
        self.wb = openpyxl.load_workbook(self.file_path)
        self.cur_workbook = self.wb[workbook_name]
        self.column_name = column_name
        self.column_letter = column_letter
        # column and row indexes
        self.column_i = self.create_column_index()
        self.row_i = self.create_row_index(self.column_name)


    def create_column_index(self):
        '''
        Creates the column index.
        '''
        column_i = {}
        for i in range(1, len(self.cur_workbook['1'])+1):
            title = self.cur_workbook.cell(row=1, column=i).value
            if title is not None:
                column_i[title] = i
        return column_i


    def create_row_index(self, column_name):
        '''
        Creates the row index.
        '''
        row_i = {}
        for i in range(1, len(self.cur_workbook[self.column_letter])):
            title = self.cur_workbook.cell(row=i+1, column=self.column_i[column_name]).value
            if title is not None:
                row_i[title] = i+1
        return row_i


    def format_cells(self, game_name, do_not_center_list=[], do_not_border_list=[]):
        '''
        Aligns specific columns to center and adds border to cells.
        '''
        align = openpyxl.styles.alignment.Alignment(
            horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=True, indent=0)
        border = openpyxl.styles.borders.Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'),
            diagonal=None, outline=True, start=None, end=None)
        for cell in self.column_i.keys():
            if cell not in do_not_center_list:
                self.cur_workbook.cell(row=self.row_i[game_name], column=self.column_i[cell]).alignment = align
            if cell not in do_not_border_list:
                self.cur_workbook.cell(row=self.row_i[game_name], column=self.column_i[cell]).border = border


    def get_cell(self, row_value, column_value):
        '''
        Gets the cell value based on the row and column
        '''
        if type(row_value) == str:
            value = str(self.cur_workbook.cell(row=self.row_i[row_value], column=self.column_i[column_value]).value)
            return value
        else:
            value = str(self.cur_workbook.cell(row=row_value, column=self.column_i[column_value]).value)
            return value


    def update_cell(self, row_value, column_value, string):
        '''
        Updates the given cell based on row and column to the given value.
        if row_value is not a string, it will be considered an exact index instead.
        '''
        if type(row_value) == str:
            self.cur_workbook.cell(row=self.row_i[row_value], column=self.column_i[column_value]).value = string
        else:
            self.cur_workbook.cell(row=row_value, column=self.column_i[column_value]).value = string
        self.changes_made = 1


    def add_new_cell(self, cell_dict):
        '''
        Adds the given dictionary onto a new line within the excel sheet.
        If dictionary keys match existing columns within the set sheet, it will add the value to that column.
        '''
        append_list = []
        for column in self.column_i:
            if column in cell_dict:
                append_list.append(cell_dict[column])
            else:
                append_list.append('')
                print(f'Missing data for {column}.')
        self.cur_workbook.append(append_list)
        self.changes_made = 1


    def save_excel_sheet(self):
        '''
        Backs up the excel file before saving the changes.
        It will keep trying to save until it completes in case of permission errors caused by the file being open.
        '''
        try:
            # backups the file before saving.
            shutil.copy(self.file_path, os.path.join(self.script_dir, self.excel_filename + '.bak'))
            # saves the file once it is closed
            print('\nSaving...')
            first_run = 1
            while True:
                try:
                    self.wb.save(self.excel_filename + '.xlsx')
                    print('Save Complete.                                  ')
                    break
                except PermissionError:
                    if first_run:
                        print('Make sure the excel sheet is closed.', end='\r')
                        first_run = 0
                    sleep(.1)
        except KeyboardInterrupt:
            print('\nCancelling Save')
            exit()


class Tracker:

    # config init
    with open('config.json') as file:
        data = json.load(file)
    steam_id = str(data['settings']['steam_id'])
    excel_filename = data['settings']['excel_filename']
    # var init
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    date_format = '%m/%d/%Y'
    # Indexer init
    excel = Indexer(
        excel_filename=excel_filename,
        workbook_name='Games',
        column_name='Game Name',
        column_letter='B')
    do_not_center_list = ['Game Name']
    do_not_border_list = []


    def get_api_key(self):
        '''
        Checks for an api_key.txt so it can retrieve the key. If it does not exists,
        it will ask for an API key so it can create an api_key.txt file.
        '''
        if os.path.isfile('api_key.txt'):
            with open('api_key.txt') as f:
                return f.read()
        else:
            api_key = ''
            with open(os.path.join(self.script_dir, 'api_key.txt'), 'w') as f:
                while len(api_key) != 32:
                    api_key = input('Enter your Steam API Key.\n')
                f.write(api_key)
            return api_key


    @staticmethod
    def hours_played(playtime_forever):
        '''
        Converts minutes played to a hours played in decimal form.
        '''
        return round(playtime_forever/60, 4)


    def get_time_to_beat(self, game_name, delay=2):
        '''
        Uses howlongtobeatpy to get the time to beat for entered game.
        '''
        if delay > 0:
            sleep(delay)
        results = HowLongToBeat().search(game_name)
        if results is not None and len(results) > 0:
            best_element = max(results, key=lambda element: element.similarity)
            time_to_beat = float(str(best_element.gameplay_main).replace('½','.5'))
            time_to_beat_unit = best_element.gameplay_main_unit
            if time_to_beat_unit == None:
                return 'No Data'
            elif time_to_beat_unit != 'Hours':
                return round(time_to_beat/60, 1)  # converts minutes to hours
            else:
                return time_to_beat
        else:
            return 'Not Found'


    def get_metacritic_score(self, game_name, platform, delay=1):
        '''
        Uses requests to get the metacritic review score for the entered game.
        '''
        sleep(delay)
        if platform == 'PS4':
            platform = 'playstation-4'
        elif platform == 'PS5':
            platform = 'playstation-5'
        elif platform in ['Steam', 'Uplay', 'Origin']:
            platform = 'pc'
        replace_dict = {
            ':':'',
            "'":'',
            '&':'',
            ',':'',
            '?':'',
            '™':'',
            '_':'-',
            ' ':'-'
            }
        for string, replace in replace_dict.items():
            game_name = game_name.replace(string, replace)
        url = f'https://www.metacritic.com/game/{platform.lower()}/{game_name.lower()}'
        user_agent = {'User-agent': 'Mozilla/5.0'}
        source = requests.get(url, headers=user_agent)
        if source.status_code == requests.codes.ok:
            soup = BeautifulSoup(source.text, 'html.parser')
            review_score = soup.find(itemprop="ratingValue")
            if review_score != None:
                review_score = int(review_score.text)
            else:
                review_score = 'No Data'
            return review_score
        else:
            # print(source.status_code)
            return 'Page Error'


    def requests_loop(self, skip_filled=1):
        '''
        Loops through games in row_i and gets missing data for time to beat and Metacritic score.
        '''
        time_to_beat_column_name = 'Time To Beat in Hours'
        metacritic_column_name = 'Metacritic Score'
        # creates checklist
        check_list = []
        for game in self.excel.row_i:
            play_status = self.excel.get_cell(game, 'Play Status')
            if play_status not in ['Unplayed', 'Playing', 'Played', 'Finished', 'Quit']:
                continue
            if skip_filled:
                time_to_beat = self.excel.get_cell(game, time_to_beat_column_name)
                metacritic_score = self.excel.get_cell(game, metacritic_column_name)
                if time_to_beat == 'None' or metacritic_score == 'None':
                    check_list.append(game)
            else:
                check_list.append(game)
        missing_data = len(check_list)
        if missing_data > 0:
            msg = f'\nSome data is missing for {missing_data} games.\nDo you want to retrieve it?\n'
            if not input(msg) in ['yes', 'y']:
                return
        else:
            return
        try:
            print('\nStarting Time To Beat and Metacritic Score check.')
            for game_name in tqdm(iterable=check_list, ascii=True, unit='games', dynamic_ncols=True):
                # How long to beat check
                time_to_beat = self.get_time_to_beat(game_name)
                if time_to_beat != None:
                    self.excel.update_cell(game_name, time_to_beat_column_name, time_to_beat)
                # metacritic score check
                platform = self.excel.get_cell(game_name, 'Platform')
                metacritic_score = self.get_metacritic_score(game_name, platform)
                if metacritic_score != None:
                    self.excel.update_cell(game_name, metacritic_column_name, metacritic_score)
        except KeyboardInterrupt:
            print('\nCancelled')
        finally:
            self.excel.save_excel_sheet()


    def refresh_steam_games(self, steam_id):
        '''
        Gets games owned by the entered Steam ID amd runs excel update/add functions.
        '''
        # asks for a steam id if the given one is invalid
        if len(steam_id) != 17:
            steam_id = input('\nInvalid Steam ID (It must be 17 numbers.)\nTry Again.\n')
        root_url = 'http://api.steampowered.com/IPlayerService/GetOwnedGames/v0001/'
        url_var = f'?key={self.get_api_key()}&steamid={steam_id}'
        combinded_url = f'{root_url}{url_var}&include_played_free_games=0&format=json&include_appinfo=1'
        try:
            data = requests.get(combinded_url)
        except requests.exceptions.ConnectionError:
            print('Internet Error')
            return False
        if data.status_code == requests.codes.ok:
            self.removed_from_steams = []
            # checks for games that changed names
            # TODO switch to using appid
            for game in self.excel.row_i.keys():
                if self.excel.get_cell(game, 'Platform') == 'Steam':
                    self.removed_from_steams.append(str(game))
            self.total_games_updated = 0
            self.total_games_added = 0
            self.added_games = []
            for game in data.json()['response']['games']:
                game_name = game['name']
                playtime_forever = game['playtime_forever']
                game_appid = game['appid']
                # Demo
                if re.search(r'\bdemo\b', game['name'].lower()) or re.search(r'\btest\b', game['name'].lower()):
                    play_status = 'Demo'  # sets play_status to Demo if demo or test appears in the name
                # Unplayed
                elif playtime_forever <= 20:
                    play_status = 'Unplayed'  # sets play_status to Unplayed if playtime_forever is 0-15
                # Played
                elif game['playtime_windows_forever'] > 10 * 60:  # hours multiplied by minutes in an hour
                    play_status = 'Played'  # sets play_status to Played if playtime_forever is greater then 10 hours
                # Unset
                else:
                    play_status = 'Unset'  # sets play_status to Unset if none of the above applies
                # Updates game if it is in the index or adds if it is not.
                if game_name in self.excel.row_i.keys():
                    try:
                        self.removed_from_steams.remove(game_name)
                    except ValueError:
                        # This is for ignoring duplicates that should not exist.
                        pass
                    self.update_game(game_name, playtime_forever, play_status)
                else:
                    self.add_game(game_name, playtime_forever, game_appid, play_status)
                    self.added_games.append(game_name)
                # formats the added/updated cells to be sure it is set right
                self.excel.format_cells(game['name'], do_not_center_list=self.do_not_center_list)
            if len(self.removed_from_steams) > 0:
                print(f'\nThe following Steam games are unaccounted for:\n{" ,".join(self.removed_from_steams)}')
            return True
        if data.status_code == 500:
            print('Server Error: make sure your api key and steam id is valid.')
        else:
            print('\nFailed to connect to Steam API:\nCheck your internet.')
            print(data.status_code)
            return False


    def update_game(self, game_name, playtime_forever, play_status):
        '''
        Updates the games playtime(if changed) and play status(if unset).
        '''
        previous_hours_played = self.excel.get_cell(game_name, 'Hours Played')
        current_hours_played = self.hours_played(playtime_forever)
        current_platform = self.excel.get_cell(game_name, 'Platform')
        if current_platform != 'Steam':  # prevents updating games that are owned on Steam and a console.
            return
        elif previous_hours_played == None or previous_hours_played == 'None':
            previous_hours_played = 0
        else:
            previous_hours_played = float(previous_hours_played)
        if current_hours_played > previous_hours_played:
            print(game_name, current_hours_played, previous_hours_played)
            self.excel.update_cell(game_name, 'Hours Played', self.hours_played(playtime_forever))
            self.excel.update_cell(game_name, 'Last Updated', dt.datetime.now().strftime(self.date_format))
            self.total_games_updated += 1
            if play_status == 'unset':
                self.excel.update_cell(game_name, 'Play Status', 'Played')
            unit = 'hours'
            added_time = current_hours_played - previous_hours_played
            if added_time < 1:
                added_time = added_time * 60
                unit = 'minutes'
            added_time = round(added_time, 1)
            print(f'\n > {game_name} updated.\n   Added {added_time} {unit}.')


    def add_game(self, game_name, playtime_forever, game_appid, play_status):
        '''
        Appends new game to excel sheet into the correct columns using self.column_i.
        Any columns that are inputted manually are left blank.
        '''
        if 'VR' in game_name.lower():
            vr_support = 'Yes'
        else:
            vr_support = ''
        time_to_beat = self.get_time_to_beat(game_name)
        metacritic_score = self.get_metacritic_score(game_name, 'Steam')
        column_info = {
            'My Rating':'',
            'Game Name':game_name,
            'Play Status':play_status,
            'Platform':'Steam',
            'VR Support':vr_support,
            'Time To Beat in Hours':time_to_beat,
            'Metacritic Score':metacritic_score,
            'Rating Comparison':'',
            # TODO add working formula that uses correct letters
            # 'Probable Completion':f'=IFERROR((G{1018}/60)/F{1018},0)',
            'Probable Completion':'',
            'Hours Played':self.hours_played(playtime_forever),
            'App ID':game_appid,
            'Last Updated':dt.datetime.now().strftime(self.date_format),
            'Date Added':dt.datetime.now().strftime(self.date_format),
            }
        self.excel.add_new_cell(column_info)
        self.total_games_added += 1
        # adds game to row_i
        self.excel.row_i[game_name] = self.excel.cur_workbook._current_row


    def output_completion_data(self):
        '''
        Shows total games added and updated games with info.
        '''
        if self.total_games_added > 0:
            print(f'\nGames Added: {self.total_games_added}')
            if len(self.added_games) > 0:
                print(', '.join(self.added_games))
        if self.total_games_updated > 0:
            print(f'\nGames Updated: {self.total_games_updated}')
        if self.excel.changes_made:
            self.excel.save_excel_sheet()
        else:
            print('\nNo games were added or updated.')


    def pick_random_game(self):
        '''
        Allows you to pick a play_status to have a random game chosen from. It allows retrying.
        '''
        print('\nWhat play status do you want a random game picked from?\nPress Enter to skip.')
        play_status_choices = {
            '1':'Played', '2':'Playing', '3':'Waiting', '4':'Finished',
            '5':'Quit', '6':'Unplayed', '7':'Ignore', '8':'Demo'
        }
        play_status = input(', '.join(play_status_choices.values()) + '\n').lower()
        if play_status == 'in':
            if len(play_status) == 1:
                play_status = play_status_choices[play_status]
            choice_list = []
            for game, index in self.excel.row_i.items():
                game_play_status = self.excel.get_cell(index, 'Play Status').lower()
                if game_play_status == play_status.lower():
                    choice_list.append(game)
            picked_game = random.choice(choice_list)
            # TODO add improvements to output
            print(f'\nPicked game with {play_status} status:\n{picked_game}')
            # allows getting another random pick
            while not input('Press Enter to pick another and No for finish.\n').lower() in ['no', 'n']:
                picked_game = random.choice(choice_list)
                print(f'\nPicked game with {play_status} status:\n{picked_game}')


    def arg_func(self):
        '''
        ph
        '''
        arg = sys.argv[1].lower()
        if arg == 'help':
            print('Help:\nrefresh- refreshes steam games\nrandom- allows getting random picks from games based on play status')
        elif arg == 'refresh':
            print('Running refresh')
            self.refresh_steam_games(self.steam_id)
        elif arg == 'random':
            self.pick_random_game()
        else:
            print('Invalid Argument Given.')


    def run(self):
        '''
        Main run function.
        '''
        # checks for arguements
        if len(sys.argv) > 1:
            self.arg_func()
        else:
            os.system('mode con cols=68 lines=40')
            print('Starting Game Tracker')
            if self.refresh_steam_games(self.steam_id):
                self.output_completion_data()
            try:
                self.requests_loop()
                self.pick_random_game()
                input('\nPress Enter to open updated file in Excel.\n')
                os.startfile(self.excel.file_path)  # opens excel file if previous input is passed
            except KeyboardInterrupt:
                print('\nClosing')


if __name__ == "__main__":
    Tracker().run()
