from Check_for_API_File import Check_for_API
from Excel_Indexer import *
import datetime as dt
import openpyxl
from openpyxl.styles.borders import Side
import requests
import shutil
import time
import os

class Tracker:

    api_key = Check_for_API()
    file_title = 'game_data'
    file_path = os.path.join(os.getcwd(), file_title + '.xlsx')
    wb = openpyxl.load_workbook(file_path)
    games = wb['Games']
    column_index, row_index = Create_Column_Row_Index(workbook=games, column_name='Game Name', column_letter='A')


    def __init__(self, steam_id):
        self.steam_id = str(steam_id)


    @staticmethod
    def playtime_conv(playtime_forever):
        '''
        Converts minutes to a written string of minutes(1) or hours(1.0).
        '''
        if playtime_forever < 61:
            playtime_converted = f'{playtime_forever} minutes'
        else:
            playtime_converted = f'{round(playtime_forever/60, 1)} hours'
        return playtime_converted


    def format_cells(self, game_name):
        '''
        Aligns and adds border to the games cells.
        '''
        # alignment setting
        alignment = openpyxl.styles.alignment.Alignment(
            horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=True, indent=0)
        center_list = [
            'My Rating', 'Play Status', 'Platform', 'VR Support', 'Release', 'Minutes Played',
            'Converted Time Played', 'App ID', 'Last Updated', 'Date Added']
        for cell in center_list:
            self.games.cell(
                row=self.row_index[game_name], column=self.column_index[cell]).alignment = alignment
        # border setting
        border = openpyxl.styles.borders.Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'),
            diagonal=None, outline=True, start=None, end=None)
        border_list = [
            'My Rating', 'Game Name', 'Play Status', 'Platform', 'VR Support', 'Release', 'Minutes Played',
            'Converted Time Played', 'App ID', 'Last Updated', 'Date Added']
        for cell in border_list:
            self.games.cell(
                row=self.row_index[game_name], column=self.column_index[cell]).border = border


    def refresh_steam_games(self):
        '''
        Gets games owned by the entered Steam ID amd runs excel update functions.
        '''
        if len(self.steam_id) != 17:
            self.steam_id = input('Invalid Steam ID (It must be 17 numbers.)\nTry Again.\n')
        root_url = 'http://api.steampowered.com/IPlayerService/GetOwnedGames/v0001/'
        url_var = f'?key={self.api_key}&steamid={self.steam_id}'
        combinded_url = f'{root_url}{url_var}&include_played_free_games=0&format=json&include_appinfo=1'
        data = requests.get(combinded_url).json()
        self.games_updated = 0
        self.games_added = 0
        added_games = []
        for item in data['response']['games']:
            game_name = item['name']
            playtime_forever = item['playtime_forever']
            game_appid = item['appid']
            if playtime_forever == 0:
                play_status = 'Unplayed'
            else:
                play_status = 'Unset'
            if game_name in self.row_index.keys():
                self.update_game(game_name, playtime_forever, play_status)
            else:
                self.add_steam_game(game_name, playtime_forever, game_appid, play_status)
                added_games.append(game_name)
            self.format_cells(item['name'])
        print(f'\nGames Added: {self.games_added}')
        if len(added_games) > 0:
            added_games_string = ', '.join(added_games)
            print(added_games_string)
        print(f'\nGames Updated: {self.games_updated}')
        # backs up the excel file before updating.
        shutil.copy(self.file_path, os.path.join(os.getcwd(), self.file_title + '.bak'))
        Complete = False
        while Complete != True:
            try:
                self.wb.save(self.file_title + '.xlsx')
                print('\nSave Complete')
                Complete = True
            except PermissionError:
                print("Can't Save Excel File. File is already open.", end='\r')
                time.sleep(1000)


    def update_game(self, game_name, playtime_forever, play_status):
        '''
        Updates the games playtime(if changed) and play status(if unset).
        '''
        current_playtime = self.games.cell(row=self.row_index[game_name],
            column=self.column_index['Minutes Played']).value
        current_platform = self.games.cell(row=self.row_index[game_name], column=self.column_index['Platform']).value
        if current_playtime == 'Minutes Played' or current_platform != 'Steam':
            return
        elif current_playtime == None:
            print(f'Missing current_playtime for {game_name}')
            current_playtime = 0
        else:
            current_playtime = int(current_playtime)
        if playtime_forever > current_playtime:
            self.games.cell(row=self.row_index[game_name],
                column=self.column_index['Minutes Played']).value = playtime_forever
            self.games.cell(row=self.row_index[game_name],
                column=self.column_index['Converted Time Played']).value = self.playtime_conv(playtime_forever)
            self.games.cell(row=self.row_index[game_name],
                column=self.column_index['Last Updated']).value = dt.datetime.now()
            self.games_updated += 1
            if play_status == 'unset':
                self.games.cell(row=self.row_index[game_name], column=self.column_index['Play Status']).value = 'Played'
            added_time = self.playtime_conv(playtime_forever - current_playtime)
            print(f'\n{game_name} updated.\nAdded {added_time}.')


    def add_steam_game(self, game_name, playtime_forever, game_appid, play_status):
        '''
        Appends new game to excel sheet into the correct columns using self.column_index.
        Any columns that are inputted manually are left blank.
        '''
        column_info = {
            'My Rating':'',
            'Game Name':game_name,
            'Platform':'Steam',
            'Release':'',
            'VR Support':'',
            'Minutes Played':playtime_forever,
            'Converted Time Played':self.playtime_conv(playtime_forever),
            'App ID':game_appid,
            'Play Status':play_status,
            'Last Updated':dt.datetime.now().date(),
            'Date Added':dt.datetime.now().date(),
        }
        append_list = []
        for column in self.column_index:
            if column in column_info:
                append_list.append(column_info[column])
            else:
                append_list.append('')
                print(f'Missing data for {column}.')
        self.games.append(append_list)
        self.games_added += 1
        # adds game to row_index early
        self.row_index[game_name] = self.games._current_row


    def main(self):
        '''
        Main init function.
        '''
        self.refresh_steam_games()
        input('\nCheck Complete.\nPress Enter to open updated file in Excel.')
        os.startfile(self.file_path)


if __name__ == "__main__":
    App = Tracker(76561197982626192)
    App.main()
    # App.align_text()
