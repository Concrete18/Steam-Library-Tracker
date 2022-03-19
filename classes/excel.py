from openpyxl.styles import PatternFill, Border
import openpyxl
from pathlib import Path
from time import sleep
import datetime as dt
import pandas as pd
import shutil


class Excel:

    changes_made = False

    def __init__(self, excel_filename):
        """
        Allows retreiving, adding, updating, deleting and formatting cells within Excel.

        `excel_filename` is the path to the excel file.
        """
        # workbook setup
        self.file_path = Path(excel_filename)
        self.wb = openpyxl.load_workbook(self.file_path)

    def save_excel_sheet(self, use_print=True, backup=True):
        """
        Backs up the excel file before saving the changes if `backup` is True.

        It will keep trying to save until it completes in case of permission errors caused by the file being open.

        `use_print` determines if info for the saving progress will be printed.
        """
        # only saves if any changes were made
        if self.changes_made:
            try:
                # backups the file before saving.
                if backup:
                    self.file_path
                    shutil.copy(self.file_path, Path(self.file_path.name + ".bak"))
                # saves the file once it is closed
                if use_print:
                    print("\nSaving...")
                first_run = True
                while True:
                    try:
                        self.wb.save(self.file_path)
                        if use_print:
                            print(f'Save Complete.{34*" "}')
                            self.changes_made = False
                        break
                    except PermissionError:
                        if first_run:
                            if use_print:
                                print("Make sure the excel sheet is closed.", end="\r")
                            first_run = False
                        sleep(1)
            except KeyboardInterrupt:
                if use_print:
                    print("\nCancelling Save")
                exit()


class Format:
    def format_cells(self, game_name):
        """
        Cell Formatter for Game Library Tracker.
        """
        center_align = openpyxl.styles.alignment.Alignment(
            horizontal="center",
            vertical="center",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=True,
            indent=0,
        )
        left_align = openpyxl.styles.alignment.Alignment(
            horizontal="left",
            vertical="center",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=True,
            indent=0,
        )
        border = Border(
            left=openpyxl.styles.borders.Side(style="thin"),
            right=openpyxl.styles.borders.Side(style="thin"),
            top=openpyxl.styles.borders.Side(style="thin"),
            bottom=openpyxl.styles.borders.Side(style="thin"),
            diagonal=None,
            outline=True,
            start=None,
            end=None,
        )
        light_grey_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        for column in self.col_idx.keys():
            cell = self.cur_sheet.cell(
                row=self.row_idx[game_name], column=self.col_idx[column]
            )
            # Percent
            if self.list_in_string(["percent", "discount"], column):
                cell.style = "Percent"
            # currency
            elif self.list_in_string(["price", "msrp"], column):
                cell.style = "Currency"
            # 1 decimal place
            if self.list_in_string(["hours played"], column):
                cell.number_format = "#,#0.0"
            # fill
            if self.list_in_string(
                ["Rating Comparison", "Probable Completion"], column
            ):
                cell.fill = light_grey_fill
            # date
            elif self.list_in_string(["last updated", "date"], column):
                cell.number_format = "MM/DD/YYYY"
            # centering
            dont_center = [
                "Name",
                "Tags",
                "Game Name",
                "Developers",
                "Publishers",
                "Genre",
            ]
            if column not in dont_center:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            # border
            cell.border = border


class Sheet(Format):
    def __init__(self, excel_object, column_name, sheet_name=None) -> None:
        """
        Allows interacting with any one sheet within the excel_object given.

        `excel_object` Excel object created using Excel class.

        `sheet_name` Name of the sheet to use.

        `column_name` Name of the main column you intend to use for identifying rows.
        """
        self.wb = excel_object.wb
        self.excel = excel_object
        self.column_name = column_name
        self.sheet_name = sheet_name
        if sheet_name:
            self.cur_sheet = self.wb[sheet_name]
        else:
            if len(self.wb.sheetnames) > 0:
                self.cur_sheet = self.wb[self.wb.sheetnames[0]]
            else:
                raise "No sheets exist."
        self.column_name = column_name
        # column and row indexes
        self.col_idx = self.get_column_index()
        self.row_idx = self.get_row_index(self.column_name)

    def get_column_index(self):
        """
        Creates the column index.
        """
        col_index = {}
        for i in range(1, len(self.cur_sheet["1"]) + 1):
            title = self.cur_sheet.cell(row=1, column=i).value
            if title is not None:
                col_index[title] = i
        return col_index

    def get_row_index(self, col_name):
        """
        Creates the row index based on `column_name`.
        """
        row_idx = {}
        total_rows = len(self.cur_sheet["A"])
        for i in range(1, total_rows):
            title = self.cur_sheet.cell(row=i + 1, column=self.col_idx[col_name]).value
            if title is not None:
                row_idx[title] = i + 1
        return row_idx

    def list_in_string(self, list, string, lowercase=True):
        """
        Returns True if any entry in the given `list` is in the given `string`.

        Setting `lowercase` to True allows you to make the check set all to lowercase.
        """
        if lowercase:
            return any(x.lower() in string.lower() for x in list)
        else:
            return any(x in string for x in list)

    @staticmethod
    def create_excel_date(datetime=None, date=True, time=True):
        """
        creates an excel date from the givin `datetime` object using =DATE().

        Defaults to the current date and time if no datetime object is given.
        """
        if datetime == None:
            datetime = dt.datetime.now()
        year = datetime.year
        month = datetime.month
        day = datetime.day
        hour = datetime.hour
        minute = datetime.minute
        if date and time:
            return f"=DATE({year}, {month}, {day})+TIME({hour},{minute},0)"
        elif date:
            return f"=DATE({year}, {month}, {day})"
        elif time:
            return f"=TIME({hour},{minute},0)"
        else:
            return None

    def get_row_col_index(self, row_value, column_value):
        """
        Gets the row and column index for the given values if they exist.

        Will return the `row_value` and `column_value` if they are numbers already.
        """
        row_key, column_key = None, None
        # row key setup
        if type(row_value) == str and row_value in self.row_idx:
            row_key = self.row_idx[row_value]
        elif type(row_value) == int:
            row_key = row_value
        # column key setup
        if type(column_value) == str and column_value in self.col_idx:
            column_key = self.col_idx[column_value]
        elif type(column_value) == int:
            column_key = column_value
        return row_key, column_key

    def get_row(self, column_value):
        """
        WIP Shows row contents in a list.
        """
        row = self.row_idx[column_value]
        data = list(self.cur_sheet.iter_rows()[row])
        print(data)
        return data

    def get_cell(self, row_value, column_value):
        """
        Gets the cell value based on the `row_value` and `column_value`.
        """
        row_key, column_key = self.get_row_col_index(row_value, column_value)
        # gets the value
        if row_key is not None and column_key is not None:
            return self.cur_sheet.cell(row=row_key, column=column_key).value
        else:
            return None

    def update_index(self, col_key):
        """
        Updates the current row with the `column_key` in the `row_idx` variable.
        """
        self.row_idx[col_key] = self.cur_sheet._current_row

    def update_cell(self, row_val, col_val, new_val, save=False):
        """
        Updates the cell based on `row_val` and `col_val` to `new_val`.

        Returns True if cell was updated and False if it was not updated.

        Saves after change if `save` is True.
        """
        row_key, column_key = self.get_row_col_index(row_val, col_val)
        if row_key is not None and column_key is not None:
            current_value = self.cur_sheet.cell(row=row_key, column=column_key).value
            # updates only if cell will actually be changed
            if new_val == "":
                new_val = None
            if current_value != new_val:
                self.cur_sheet.cell(row=row_key, column=column_key).value = new_val
                if save:
                    self.excel.save_excel_sheet(use_print=False, backup=False)
                else:
                    self.excel.changes_made = True
                return True
        return False

    def add_new_line(self, cell_dict, column_key, save=False, debug=False):
        """
        Adds the given dictionary, as `cell_dict`, onto a new line within the excel sheet.

        If dictionary keys match existing columns within the set sheet, it will add the value to that column.

        Use `debug` to print info if a column in the `cell_dict` does not exist.

        Saves after change if `save` is True.
        """
        append_list = []
        for column in self.col_idx:
            if column in cell_dict:
                append_list.append(cell_dict[column])
            else:
                append_list.append("")
        self.cur_sheet.append(append_list)
        self.update_index(column_key)
        if save:
            self.excel.save_excel_sheet(use_print=False, backup=False)
        else:
            self.excel.changes_made = True
        return True

    def delete_by_row(self, col_val, save=False):
        """
        Deletes row by `column_value`.
        """
        if col_val not in self.row_idx:
            return None
        row = self.row_idx[col_val]
        self.cur_sheet.delete_rows(row)
        if save:
            self.excel.save_excel_sheet(use_print=False, backup=False)
        else:
            self.excel.changes_made = True
        return True

    def delete_by_column(self, column_name):
        """
        Deletes column by `column_name`.
        """
        if column_name not in self.col_idx:
            return None
        column = self.col_idx[column_name]
        self.cur_sheet.delete_column(column)
        self.excel.changes_made = True
        return True

    def create_dataframe(self, date_columns=None, na_values=None):
        return
        # file_loc = self.excel.file_path
        # df = pd.read_excel(file_path=file_loc, engine="openpyxl", sheet_name=self.sheet_name parse_dates=date_columns, na_values=na_values)
        # return df
