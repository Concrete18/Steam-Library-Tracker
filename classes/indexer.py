from openpyxl.styles.borders import Side
from classes.logger import Logger
import openpyxl, os, shutil
from pathlib import Path
from time import sleep
import datetime as dt

class Indexer(Logger):

    changes_made = 0


    def __init__(self, excel_filename, workbook_name, column_name, script_dir):
        '''
        Allows Retreiving, adding, updating, deleting and formatting cells within Excel.

        `excel_filename` is the relative path to the excel file.

        `workbook_name` Name of the workbook that use.

        `column_name` Name of the main column you intend to use for identifying rows.
        '''
        self.script_dir = script_dir
        self.excel_filename = excel_filename
        self.file_path = os.path.join(script_dir, excel_filename + '.xlsx')
        self.wb = openpyxl.load_workbook(self.file_path)
        self.cur_workbook = self.wb[workbook_name]
        self.column_name = column_name
        # column and row indexes
        self.col_index = self.create_column_index()
        self.row_index = self.create_row_index(self.column_name)


    def create_column_index(self):
        '''
        Creates the column index.
        '''
        col_index = {}
        for i in range(1, len(self.cur_workbook['1'])+1):
            title = self.cur_workbook.cell(row=1, column=i).value
            if title is not None:
                col_index[title] = i
        return col_index

    def create_row_index(self, column_name):
        '''
        Creates the row index based on `column_name`.
        '''
        row_index = {}
        total_rows = len(self.cur_workbook['A'])
        for i in range(1, total_rows):
            title = self.cur_workbook.cell(row=i+1, column=self.col_index[column_name]).value
            if title is not None:
                row_index[title] = i+1
        return row_index

    def list_in_string(self, list, string):
        return any(x in string for x in list)
        
    def format_cells(self, game_name):
        '''
        Aligns specific columns to center and adds border to cells.
        '''
        center = openpyxl.styles.alignment.Alignment(
            horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=True, indent=0)
        left = openpyxl.styles.alignment.Alignment(
            horizontal='left', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=True, indent=0)
        border = openpyxl.styles.borders.Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'),
            diagonal=None, outline=True, start=None, end=None)
        for column in self.col_index.keys():
            cell = self.cur_workbook.cell(row=self.row_index[game_name], column=self.col_index[column])
            column = column.lower()
            # Percent
            if self.list_in_string(['percent', 'discount'], column):
                cell.style = 'Percent'
            # 1 decimal place
            if self.list_in_string(['hours played'], column):
                cell.number_format = '#,#0.0'
            # currency
            elif self.list_in_string(['price', 'msrp'], column):
                cell.style = 'Currency'
                # date
            elif self.list_in_string(['last updated', 'date'], column):
                cell.number_format = "MM/DD/YYYY"
            # centering
            if column not in ['name', 'tags', 'game name', 'developers', 'publishers', 'genre']:
                cell.alignment = center
            else:
                cell.alignment = left
            # border
            cell.border = border

    @staticmethod
    def create_excel_date(datetime=None):
        '''
        creates an excel date from the givin `datetime` object using =DATE().

        Defaults to the current date and time if no datetime object is given.
        '''
        if datetime == None:
            datetime = dt.datetime.now()
        year = datetime.year
        month = datetime.month
        day = datetime.day
        hour = datetime.hour
        minute = datetime.minute
        excel_date = f'=DATE({year}, {month}, {day})+TIME({hour},{minute},0)'
        return excel_date

    def get_row_col_index(self, row_value, column_value):
        '''
        Gets the row and column index for the given values if they exist.

        Will return the `row_value` and `column_value` if they are numbers already.
        '''
        row_key, column_key = None, None
        # row key setup
        if type(row_value) == str and row_value in self.row_index:
            row_key = self.row_index[row_value]
        elif type(row_value) == int:
            row_key = row_value
        # column key setup
        if type(column_value) == str and column_value in self.col_index:
            column_key = self.col_index[column_value]
        elif type(column_value) == int:
            column_key = column_value
        return row_key, column_key

    def get_cell(self, row_value, column_value):
        '''
        Gets the cell value based on the `row_value` and `column_value`.
        '''
        row_key, column_key = self.get_row_col_index(row_value, column_value)
        # gets the value
        if row_key is not None and column_key is not None:
            return self.cur_workbook.cell(row=row_key, column=column_key).value
        else:
            return None

    def update_cell(self, row_value, column_value, value):
        '''
        Updates the given cell based on row and column to the given value.
        if row_value is not a string, it will be considered an exact index instead.
        '''
        row_key, column_key = self.get_row_col_index(row_value, column_value)
        # updates value
        if row_key is not None and column_key is not None:
            self.cur_workbook.cell(row=row_key, column=column_key).value = value
            self.changes_made = True
        else:
            return None

    def add_new_cell(self, cell_dict):
        '''
        Adds the given dictionary onto a new line within the excel sheet.
        If dictionary keys match existing columns within the set sheet, it will add the value to that column.
        '''
        append_list = []
        for column in self.col_index:
            if column in cell_dict:
                append_list.append(cell_dict[column])
            else:
                append_list.append('')
                self.logger.info(f'Missing {column} for {cell_dict[self.column_name]}.')
        self.cur_workbook.append(append_list)
        self.changes_made = 1

    def save_excel_sheet(self, show_print=True):
        '''
        Backs up the excel file before saving the changes.
        It will keep trying to save until it completes in case of permission errors caused by the file being open.
        '''
        try:
            # backups the file before saving.
            shutil.copy(self.file_path, os.path.join(self.script_dir, self.excel_filename + '.bak'))
            # saves the file once it is closed
            if show_print:
                print('\nSaving...')
            first_run = True
            while True:
                try:
                    self.wb.save(self.excel_filename + '.xlsx')
                    if show_print:
                        print('Save Complete.                                  ')
                        self.changes_made = False
                    break
                except PermissionError:
                    if first_run:
                        if show_print:
                            print('Make sure the excel sheet is closed.', end='\r')
                        first_run = False
                    sleep(.1)
        except KeyboardInterrupt:
            if show_print:
                print('\nCancelling Save')
            exit()
