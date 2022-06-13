from classes.excel import Sheet
from openpyxl.styles import PatternFill, Border
import openpyxl


class CustomSheet(Sheet):
    def format_cells(self, game_name):
        """
        Cell Formatter for Game Library Tracker.
        """
        center_align = openpyxl.styles.alignment.Alignment(
            horizontal="center",
            vertical="center",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=True,
            indent=0,
        )
        left_align = openpyxl.styles.alignment.Alignment(
            horizontal="left",
            vertical="center",
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=True,
            indent=0,
        )
        border = Border(
            left=openpyxl.styles.borders.Side(style="thin"),
            right=openpyxl.styles.borders.Side(style="thin"),
            top=openpyxl.styles.borders.Side(style="thin"),
            bottom=openpyxl.styles.borders.Side(style="thin"),
            diagonal=None,
            outline=True,
            start=None,
            end=None,
        )
        light_grey_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        for column in self.col_idx.keys():
            row_i = self.row_idx[game_name]
            col_i = self.col_idx[column]
            cell = self.cur_sheet.cell(row_i, col_i)
            # Percent
            if self.list_in_string(
                [
                    "percent",
                    "discount",
                    "Rating Comparison",
                    "Probable Completion",
                ],
                column,
            ):
                cell.style = "Percent"
            # interger
            elif self.list_in_string(
                [
                    "Days Since Updated",
                ],
                column,
            ):
                cell.number_format = "0"
            # 1 decimal place
            elif self.list_in_string(
                [
                    "hours played",
                    # "Days Since Updated",
                ],
                column,
            ):
                cell.number_format = "#,#0.0"
            # currency
            elif self.list_in_string(["price", "msrp"], column):
                cell.style = "Currency"
            # date
            elif self.list_in_string(["last updated", "date"], column):
                cell.number_format = "MM/DD/YYYY"
            # fill
            if self.list_in_string(
                ["Rating Comparison", "Probable Completion"], column
            ):
                cell.fill = light_grey_fill
            # centering
            dont_center = [
                "Name",
                "Tags",
                "Game Name",
                "Developers",
                "Publishers",
                "Genre",
            ]
            if column not in dont_center:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            # border
            cell.border = border
